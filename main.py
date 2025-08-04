import openpyxl
from random import shuffle
from pathlib import Path
import tkinter as tk
from tkinter import filedialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *


def ensure_output_folder():
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    return output_dir


def validate_file(file_path):
    if not file_path.endswith('.xlsx'):
        raise ValueError("Файл должен быть в формате .xlsx")
    if not Path(file_path).exists():
        raise FileNotFoundError("Файл не найден")
    return file_path


def copy_cell_style(source_cell, target_cell):
    if source_cell.font:
        target_cell.font = openpyxl.styles.Font(
            name=source_cell.font.name, size=source_cell.font.size,
            bold=source_cell.font.bold, italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign, underline=source_cell.font.underline,
            strike=source_cell.font.strike, color=source_cell.font.color
        )
    else:
        target_cell.font = None
    if source_cell.border:
        target_cell.border = openpyxl.styles.Border(
            left=source_cell.border.left, right=source_cell.border.right,
            top=source_cell.border.top, bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal, diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline
        )
    else:
        target_cell.border = None
    if source_cell.fill:
        target_cell.fill = openpyxl.styles.PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color, end_color=source_cell.fill.end_color
        )
    else:
        target_cell.fill = None
    target_cell.number_format = source_cell.number_format
    if source_cell.protection:
        target_cell.protection = openpyxl.styles.Protection(
            locked=source_cell.protection.locked, hidden=source_cell.protection.hidden
        )
    else:
        target_cell.protection = None
    if source_cell.alignment:
        target_cell.alignment = openpyxl.styles.Alignment(
            horizontal=source_cell.alignment.horizontal, vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation, wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit, indent=source_cell.alignment.indent
        )
    else:
        target_cell.alignment = None


def shuffle_excel_rows(file_path, start_row, end_row, start_col, end_col):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        max_rows, max_cols = ws.max_row, ws.max_column

        if start_row < 1 or end_row < start_row or end_row > max_rows or \
           start_col < 1 or end_col < start_col or end_col > max_cols:
            raise ValueError(f"Неверный диапазон. Строки: 1-{max_rows}, Столбцы: 1-{max_cols}")

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append((cell.value, cell))
            data.append(row_data)

        shuffle(data)

        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, (value, source_cell) in enumerate(row_data, start=start_col):
                target_cell = ws.cell(row=row_idx, column=col_idx)
                target_cell.value = value
                copy_cell_style(source_cell, target_cell)

        for row_idx in range(start_row, end_row + 1):
            ws.row_dimensions[row_idx].height = None

        output_dir = ensure_output_folder()
        output_file = output_dir / f"{Path(file_path).stem}_shuffled.xlsx"
        wb.save(output_file)
        return f"Перемешанный файл сохранен как {output_file}"
    except Exception as e:
        return f"Ошибка: {str(e)}"


class ExcelShufflerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Row Shuffler")
        self.root.geometry("500x400")

        # Main frame with padding
        main_frame = ttk.Frame(root, padding=20)
        main_frame.pack(fill=BOTH, expand=True)

        # Title
        ttk.Label(main_frame, text="Перемешиватель строк Excel", font=("Arial", 16, "bold")).pack(pady=10)

        # File selection frame
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill=X, pady=5)
        self.file_path = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path, width=30).pack(side=LEFT, padx=5)
        ttk.Button(file_frame, text="Обзор", command=self.browse_file, bootstyle=SECONDARY).pack(side=LEFT)

        # Row range
        ttk.Label(main_frame, text="Диапазон строк (например, 1-10):").pack(anchor=W, pady=5)
        self.row_range = ttk.Entry(main_frame, width=20)
        self.row_range.pack(anchor=W, padx=5)

        # Column range
        ttk.Label(main_frame, text="Диапазон столбцов (например, 1-5):").pack(anchor=W, pady=5)
        self.col_range = ttk.Entry(main_frame, width=20)
        self.col_range.pack(anchor=W, padx=5)

        # Shuffle button
        ttk.Button(main_frame, text="Перемешать строки", command=self.run_shuffle,
                  bootstyle="primary", width=20, style="large.TButton").pack(pady=20)

        # Status label
        self.status = ttk.Label(main_frame, text="", wraplength=450, font=("Arial", 10))
        self.status.pack(fill=X, pady=10)

        # Custom style for shuffle button
        style = ttk.Style()
        style.configure("large.TButton", font=("Arial", 12))


    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Файлы Excel", "*.xlsx")])
        if file_path:
            self.file_path.set(file_path)


    def run_shuffle(self):
        try:
            file_path = self.file_path.get()
            if not file_path:
                raise ValueError("Выберите файл")
            start_row, end_row = map(int, self.row_range.get().split('-'))
            start_col, end_col = map(int, self.col_range.get().split('-'))
            result = shuffle_excel_rows(file_path, start_row, end_row, start_col, end_col)
            self.status.config(text=result, bootstyle=SUCCESS)
        except Exception as e:
            self.status.config(text=f"Ошибка: {str(e)}", bootstyle=DANGER)


if __name__ == "__main__":
    root = ttk.Window(themename="flatly")  # Windows 11-like theme
    app = ExcelShufflerApp(root)
    root.mainloop()
